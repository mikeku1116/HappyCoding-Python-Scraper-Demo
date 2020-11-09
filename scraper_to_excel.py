from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font

response = requests.get("https://www.inside.com.tw/tag/AI")

soup = BeautifulSoup(response.content, "lxml")

cards = soup.find_all("div", {"class": "post_list_item"})

results = []
for card in cards:
    title = card.find("h3", {"class": "post_title"})
    published = card.find("li", {"class": "post_date"})

    results.append((title.getText().strip(),) + (published.getText().strip(),))

wb = openpyxl.Workbook()  # 建立工作簿
sheet = wb.create_sheet("inside", 0)  # 建立工作表

sheet.append(("文章標題", "發佈日期"))  # 寫入欄位名稱
sheet.cell(row=1, column=1).font = Font(color="0000FF")  # 顯示藍色字體顏色
sheet.cell(row=1, column=2).font = Font(color="0000FF")  # 顯示藍色字體顏色

for result in results:
    sheet.append(result)  # 寫入爬取的資料

wb.save("inside.xlsx")  # 儲存Excel檔案
